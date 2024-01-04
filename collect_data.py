

# Python program to read an excel file

# import openpyxl module
import openpyxl
from datetime import date


def calculateDay(from_date, to_date):
    # format date is dd/mm/yyyy
    from_date = from_date.split("/")
    to_date = to_date.split("/")
    from_date = date(int(from_date[2]), int(from_date[1]), int(from_date[0]))
    to_date = date(int(to_date[2]), int(to_date[1]), int(to_date[0]))
    return (to_date - from_date).days


# Give the location of the file
path = "data/data.xlsm"
to_day = "1/9/2023"  # DD/MM/YYYY

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path, data_only=True)

# print(wb_obj.sheetnames)
# Get workbook active sheet object
# from the active attribute
cost_sheet = wb_obj["Cost DRP(2)"]

cost_data = []

for row in cost_sheet.iter_rows(min_row=3, max_row=23, min_col=1, max_col=6, values_only=True):
    is_empty = True
    rows = []
    for cell in row:
        if cell is not None:
            is_empty = False
            rows.append(cell)
    if not is_empty:
        cost_data.append(rows)

stock_sheet = wb_obj["Stock @ 1SEP"]
stock_data = []

for row in stock_sheet.iter_rows(min_row=2, max_row=1000, min_col=1, max_col=6, values_only=True):
    is_empty = True
    rows = []
    for cell in row:
        if cell is not None:
            is_empty = False
            rows.append(cell)
    if not is_empty:
        stock_data.append(rows)


demand_sheet = wb_obj["Demand (2)"]
demand_data = []

for row in demand_sheet.iter_rows(min_row=1, max_row=1000, min_col=2, max_col=19, values_only=True):
    rows = []
    for cell in row:
        rows.append(cell)
    if any(rows):
        demand_data.append(rows)
    else:
        break


def check_C_name_of_city(city_name):
    for row in cost_data:
        if row[1] == city_name:
            return row[3]


temp = []
for row in demand_data:
    row = row[0:3]+[check_C_name_of_city(row[2])] + row[3:]
    temp.append(row)

demand_data = temp

demand_filter = []
for row in demand_data:
    for ind, cell in enumerate(row):
        if ind >= 3 and ind < len(row)-1 and cell is not None:
            if demand_data[0][ind] is not None and row[0] is not None and row[1] is not None and row[3] is not None and cell is not None:
                if int(cell) > 0:
                    demand_filter.append({
                        "day": demand_data[0][ind],
                        'material': row[0],
                        # "from": row[1],
                        "to": row[3],
                        "demand": cell,
                        "deadline": calculateDay(to_day, demand_data[0][ind])
                    })


def standardization_data(G, data):
    demand_date = data.get("day")
    start = G.get_node(data.get("to"))  # comvert name to C
    demand = int(data.get("demand"))
    material = int(data.get("material"))
    deadline = int(data.get("deadline"))

    return demand_date, start, demand, material, deadline


# step:
# 1: lấy data từ file excel
# 2: filter data, lấy giá trị cần
# 3: chuẩn hoá data
